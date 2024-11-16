#!/usr/bin/env python
# coding: utf-8

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

class TrainSimulationReport:
    def __init__(
                self,
                excel_path: str,
                plot_path: str,
                df_speed: pd.core.frame.DataFrame,
                station_names,
                distances_between_stations,
                section_avg_speed_log,
                messages_string
            ):
        self.excel_path = excel_path
        self.plot_path = plot_path
        self.df_speed = df_speed
        self.station_names = station_names
        self.distances_between_stations = distances_between_stations
        self.section_avg_speed_log = section_avg_speed_log
        self.messages_string = messages_string

    def generate_spreadsheet(self) -> None:
        """
        Generates a spreadsheet report of the train simulation, including a plot.
        """
        # Save data to the Excel file
        self.df_speed.to_excel(self.excel_path, index=False, sheet_name="Data")

        # Create a DataFrame for chainage and section average speed
        df_average = self._chainage(self.station_names, self.distances_between_stations, self.section_avg_speed_log)

        # Append the new DataFrame to a new sheet in the same Excel file
        with pd.ExcelWriter(self.excel_path, mode='a', engine='openpyxl') as writer:
            df_average.to_excel(writer, index=False, sheet_name="Speed")

        # Insert the plot image into the Excel file
        wb = load_workbook(self.excel_path)
        ws_plot = wb.create_sheet("Plot")
        img = Image(self.plot_path)
        ws_plot.add_image(img, "A1")

        # Insert custom text at the top of the "Speed" sheet
        ws_speed = wb["Speed"]
        for _ in range(6):
            ws_speed.insert_rows(1)
        # Write each message to consecutive cells in column A
        for i, message in enumerate(self.messages_string, start=1):
            cell = f"A{i}"  # Generate cell reference like A1, A2, etc.
            ws_speed[cell] = message

        wb.save(self.excel_path)

    @staticmethod
    def _chainage(station_names, segment_distance_log, section_avg_speed_log):
        """
        Creates a DataFrame for station chainages and section average speeds.
        """
        # Calculate distances between consecutive stations
        from_station = []
        to_station = []
        interstation_distance = []
        seperator = []
        sl_no = []
        for i in range(len(segment_distance_log)-1):
            sl_no.append(i+1)
            from_station.append(station_names[i])
            to_station.append(station_names[i+1])
            interstation_distance.append(int(segment_distance_log[i+1]))
            seperator.append('--')
    
        data = {
            "Sl.no": sl_no,
            "Station From":from_station,
            " ":seperator,
            "Station To":to_station,
            "Distance":interstation_distance,
            "Section Speed":[float(f'{float(i):.2f}') for i in section_avg_speed_log]
        }
   
        # Create DataFrame
        return pd.DataFrame(data)
    
 